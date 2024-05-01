import os
import time
from telethon.sync import TelegramClient
from telethon.tl.functions.messages import GetDialogsRequest
from telethon.tl.types import InputPeerEmpty, Chat, Channel
import openpyxl

def channelandgroups(api_id, api_hash, print_pages, print_channel_lists):
    os.system('cls||clear')
    chats = []
    last_date = None
    size_chats = 500
    groups = []
    exit_flag = False
    openchannels = []
    closechannels = []
    openchats = []
    closechats = []

    while not exit_flag:
        os.system('cls||clear')
        sessions = [file for file in os.listdir('.') if file.endswith('.session')]

        for i in range(len(sessions)):
            print(f"[{i}] - {sessions[i]}")
        print()

        user_input = input("Выберите существующий аккаунт для получения ссылок на подключенные чаты (e - назад): ")
        if user_input.lower() == 'e':
            break
        else:
            try:
                i = int(user_input)
                if 0 <= i < len(sessions):
                    client = TelegramClient(sessions[i].replace('\n', ''), api_id, api_hash)
                    client.connect()

                    #qqqs = client.get_dialogs()

                    #for qqq in qqqs:
                        #print(qqq)
                    #input("нажми")
                    #break
                    
                    # Получение информации о пользователе
                    me = client.get_me()
                    userid = me.id
                    firstname = me.first_name
                    username = f"@{me.username}" if me.username is not None else ""
                    lastname = me.last_name if me.last_name is not None else ""
                    phone = sessions[i].split('.')[0]

                    chats = client.get_dialogs()
                    for chat in chats:
                        if isinstance(chat.entity, Channel) or isinstance(chat.entity, Chat): #проверяем групповой ли чат
                            # Определяем открытый канал
                            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast') and chat.entity.participants_count != None:
                                if chat.entity.broadcast and chat.entity.username:
                                    openchannels.append(chat.entity)
                                    groups.append(chat.entity)
                                    
                            # Определяем закрытый канал
                            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                                if chat.entity.broadcast and chat.entity.username == None and chat.entity.title != 'Unsupported Chat':
                                    closechannels.append(chat.entity)
                                    groups.append(chat.entity)
                                    
                            # Определяем открытый чат
                            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                                if chat.entity.broadcast == False and chat.entity.username:
                                    openchats.append(chat.entity)
                                    groups.append(chat.entity)
                            groups.append(chat.entity)
                            
                            # Определяем закрытый чат
                            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                                if chat.entity.broadcast == False and chat.entity.username == None:
                                    closechats.append(chat.entity)
                                    groups.append(chat.entity)
                            if isinstance(chat.entity, Chat) and chat.entity.migrated_to is None:
                                closechats.append(chat.entity)
                                groups.append(chat.entity)                            
                    
        
                    while True:
                        os.system('cls||clear')
                        oc = 0
                        cc = 0
                        og = 0
                        cg = 0
                        owner_channel = 0
                        owner_group = 0
                        owner_closegroup = 0
                        owner_closechannel = 0
                        print('-----------------------------')
                        print("=ИНФОРМАЦИЯ О КАНАЛАХ И ЧАТАХ=")
                        print(f"\033[96mНомер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}\033[0m")
                        print('-----------------------------')
                        print()

                        
                        # Код для формирования списков openchannels, closechannels, openchats и closechats...
                        print_channel_lists(openchannels, closechannels, openchats, closechats, print_pages)


                        print()
                        print("---------------------------------------")
                        print(f"Открытые каналы: {oc}")
                        print(f"Открытые группы: {og}")
                        print()
                        print(f"\033[91mЗакрытые каналы: {cc}\033[0m")
                        print(f"\033[91mЗакрытые группы: {cg}\033[0m")
                        print("---------------------------------------")
                        print(f"\033[96mИмеет права владельца или админа в {owner_channel} каналах, из них {owner_closechannel} - в закрытых\033[0m")
                        print(f"\033[96mИмеет права владельца или админа в {owner_group} группах, из них {owner_closegroup} - в закрытых\033[0m")

                        #g_index_str = str(input("Для выгрузки информаци в файл Excel, введите 'get', для возврата - введеите 'e': "))
                        g_index_str = str('get')
                        print()
                        input("Для продолжение нажмите любую клавишу, информация о группах будет автоматически сохранена в файл Excel  ")

                        if g_index_str.lower() == 'e':
                            client.disconnect()
                            groups = []
                            chats = []
                            openchannels = []
                            closechannels = []
                            openchats = []
                            closechats = []
                            openchannel_list = []
                            owner_channel = 0
                            owner_group = 0
                            owner_closegroup = 0
                            owner_closechannel = 0
                            break
                        else:
                            try:
                                if g_index_str == "get":
                                    wb = openpyxl.Workbook()
                                    
                                    ws = wb.active
                                    ws.append([f"Номер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}"])
                                    ws.append([f"Открытые каналы: {oc}"])
                                    ws.append([f"Открытые группы: {og}"])
                                    ws.append([f"Закрытые каналы: {cc}"])
                                    ws.append([f"Закрытые группы: {cg}"])
                                    ws.append([f"Имеет права владельца или админа в {owner_channel} каналах, из них {owner_closechannel} - в закрытых"])
                                    ws.append([f"Имеет права владельца или админа в {owner_group} группах, из них {owner_closegroup} - в закрытых"])

                                    ws_open_channels = wb.create_sheet("Открытые Каналы")
                                    ws_closed_channels = wb.create_sheet("Закрытые Каналы")
                                    ws_open_groups = wb.create_sheet("Открытые Группы")
                                    ws_closed_groups = wb.create_sheet("Закрытые Группы")
                                    write_data(ws_open_channels, openchannels)
                                    write_data(ws_closed_channels, closechannels)
                                    write_data(ws_open_groups, openchats)
                                    write_data(ws_closed_groups, closechats)
                                    wb.save(f"{sessions[i].replace('.session', '')}_about.xlsx")
                                    os.system('cls||clear')
                                    print('Ссылки на чаты добавлены в файл, мой командир')
                                    time.sleep(3)
                                    exit_flag = True
                                    client.disconnect()
                                    break
                                else:
                                    print("Пожалуйста, сделайте свой выбор")
                                    time.sleep(2)
                            except ValueError:
                                print("Пожалуйста, сделайте свой выбор")
                                time.sleep(2)
                else:
                    print("Пожалуйста, выберите существующий аккаунт в диапазоне от 0 до", len(sessions) - 1)
                    time.sleep(2)
            except ValueError:
                print("Пожалуйста, выберите существующий аккаунт в диапазоне от 0 до", len(sessions) - 1)
                time.sleep(2)

def write_data(sheet, data):
    sheet.append(["Название", "Количество участников", "Владелец", "Администратор", "ID", "Ссылка"])
    for item in data:
        owner = " (Владелец)" if item.creator else ""
        admin = " (Администратор)" if item.admin_rights is not None else ""
        usernameadd = f"@{item.username}" if hasattr(item, 'username') and item.username is not None else ""
        sheet.append([item.title, item.participants_count, owner, admin, item.id, usernameadd])
        #sheet.append([item.title, item.participants_count, owner, admin, item.id, f"@{item.username}" if hasattr(item, 'username') else ""])
        
