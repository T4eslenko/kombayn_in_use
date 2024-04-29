import os
import time
from telethon.sync import TelegramClient
from telethon.tl.functions.messages import GetDialogsRequest
from telethon.tl.types import InputPeerEmpty, Chat, Channel
import openpyxl

def channelandgroups(api_id, api_hash):
    os.system('cls||clear')
    chats = []
    last_date = None
    size_chats = 500
    groups = []
    exit_flag = False
    openchannels = []
    closechannels = []
    openchats = []
    closechats = []

    while not exit_flag:
        os.system('cls||clear')
        sessions = [file for file in os.listdir('.') if file.endswith('.session')]

        for i in range(len(sessions)):
            print(f"[{i}] - {sessions[i]}")
        print()

        user_input = input("Выберите существующий аккаунт для получения ссылок на подключенные чаты (e - назад): ")
        if user_input.lower() == 'e':
            break
        else:
            try:
                i = int(user_input)
                if 0 <= i < len(sessions):
                    client = TelegramClient(sessions[i].replace('\n', ''), api_id, api_hash)
                    client.connect()

                    #qqqs = client.get_dialogs()

                    #for qqq in qqqs:
                        #print(qqq)
                    #input("нажми")
                    #break
                    
                    # Получение информации о пользователе
                    me = client.get_me()
                    userid = me.id
                    firstname = me.first_name
                    username = f"@{me.username}" if me.username is not None else ""
                    lastname = me.last_name if me.last_name is not None else ""
                    phone = sessions[i].split('.')[0]

                    #result = client(GetDialogsRequest(
                    #    offset_date=last_date,
                    #    offset_id=0,
                    #    offset_peer=InputPeerEmpty(),
                    #    limit=size_chats,
                    #    hash=0
                    #))
                    #chats.extend(result.chats)
                    
                    chats = client.get_dialogs()
                    for chat in chats:
                        if isinstance(chat.entity, Channel) or isinstance(chat.entity, Chat): #проверяем групповой ли чат
                            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                                if chat.entity.broadcast == False and chat.entity.username == None:
                                    closechats.append(chat.entity)
                                    groups.append(chat.entity)
                            if isinstance(chat.entity, Chat) and chat.entity.migrated_to is None:
                                closechats.append(chat.entity)
                                groups.append(chat.entity)
    
                            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast') and chat.entity.participants_count != None:
                                if chat.entity.broadcast and chat.entity.username:
                                    openchannels.append(chat.entity)
                                    groups.append(chat.entity)
    
                            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                                if chat.entity.broadcast and chat.entity.username == None and chat.entity.title != 'Unsupported Chat':
                                    closechannels.append(chat.entity)
                                    groups.append(chat.entity)
    
                            if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                                if chat.entity.broadcast == False and chat.entity.username:
                                    openchats.append(chat.entity)
                                    groups.append(chat.entity)
                            groups.append(chat.entity)
                    
        
                    while True:
                        os.system('cls||clear')
                        oc = 0
                        cc = 0
                        og = 0
                        cg = 0
                        owner_channel = 0
                        owner_group = 0
                        owner_closegroup = 0
                        owner_closechannel = 0
                        print('-----------------------------')
                        print("=ИНФОРМАЦИЯ О КАНАЛАХ И ЧАТАХ=")
                        print(f"\033[96mНомер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}\033[0m")
                        print('-----------------------------')
                        print()
                        print("\033[95mОткрытые КАНАЛЫ:\033[0m")
                        for openchannel in openchannels:
                            owner = " (Владелец)" if openchannel.creator else ""
                            admin = " (Администратор)" if openchannel.admin_rights is not None else ""
                            print(f"{oc} - {openchannel.title} \033[93m[{openchannel.participants_count}]\033[0m\033[91m {owner} {admin}\033[0m ID:{openchannel.id} \033[94m@{openchannel.username}\033[0m")
                            oc += 1
                            if owner !="" or admin != "":
                                owner_channel += 1
                        
                        print()
                        print("\033[95mЗакрытые КАНАЛЫ:\033[0m")
                        for closechannel in closechannels:
                            owner = " (Владелец)" if closechannel.creator else ""
                            admin = " (Администратор)" if closechannel.admin_rights is not None else ""
                            print(f"{cc} - {closechannel.title} \033[93m[{closechannel.participants_count}]\033[0m \033[91m{owner} {admin}\033[0m ID:{closechannel.id}")
                            cc += 1
                            if owner !="" or admin != "":
                                owner_channel += 1
                                owner_closechannel += 1
                        
                        print()
                        print("\033[95mОткрытые ГРУППЫ:\033[0m")
                        for openchat in openchats:
                            owner = " (Владелец)" if openchat.creator else ""
                            admin = " (Администратор)" if openchat.admin_rights is not None else ""
                            print(f"{og} - {openchat.title} \033[93m[{openchat.participants_count}]\033[0m\033[91m {owner} {admin}\033[0m ID:{openchat.id} \033[94m@{openchat.username}\033[0m")
                            og += 1
                            if owner !="" or admin != "":
                                owner_group += 1
                        
                        print()
                        print("\033[95mЗакрытые ГРУППЫ:\033[0m")
                        for closechat in closechats:
                            owner = " (Владелец)" if closechat.creator else ""
                            admin = " (Администратор)" if closechat.admin_rights is not None else ""
                            print(f"{cg} - {closechat.title} \033[93m[{closechat.participants_count}]\033[0m \033[91m{owner} {admin}\033[0m ID:{closechat.id}")
                            cg += 1
                            if owner !="" or admin != "":
                                owner_group += 1
                                owner_closegroup += 1
                     
                        print()
                        print("---------------------------------------")
                        print(f"Открытые каналы: {oc}")
                        print(f"Открытые группы: {og}")
                        print()
                        print(f"\033[91mЗакрытые каналы: {cc}\033[0m")
                        print(f"\033[91mЗакрытые группы: {cg}\033[0m")
                        print("---------------------------------------")
                        print(f"\033[96mИмеет права владельца или админа в {owner_channel} каналах, из них {owner_closechannel} - в закрытых\033[0m")
                        print(f"\033[96mИмеет права владельца или админа в {owner_group} группах, из них {owner_closegroup} - в закрытых\033[0m")

                        #g_index_str = str(input("Для выгрузки информаци в файл Excel, введите 'get', для возврата - введеите 'e': "))
                        g_index_str = str('get')
                        input("Для продолжение нажмите любую клавишу, информация о группах будет автоматически сохранена в файл Excel")

                        if g_index_str.lower() == 'e':
                            client.disconnect()
                            groups = []
                            chats = []
                            openchannels = []
                            closechannels = []
                            openchats = []
                            closechats = []
                            owner_channel = 0
                            owner_group = 0
                            owner_closegroup = 0
                            owner_closechannel = 0
                            break
                        else:
                            try:
                                if g_index_str == "get":
                                    wb = openpyxl.Workbook()
                                    
                                    ws = wb.active
                                    ws.append([f"Номер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}"])
                                    ws.append([f"Открытые каналы: {oc}"])
                                    ws.append([f"Открытые группы: {og}"])
                                    ws.append([f"Закрытые каналы: {cc}"])
                                    ws.append([f"Закрытые группы: {cg}"])
                                    ws.append([f"Имеет права владельца или админа в {owner_channel} каналах, из них {owner_closechannel} - в закрытых"])
                                    ws.append([f"Имеет права владельца или админа в {owner_group} группах, из них {owner_closegroup} - в закрытых"])

                                    ws_open_channels = wb.create_sheet("Открытые Каналы")
                                    ws_closed_channels = wb.create_sheet("Закрытые Каналы")
                                    ws_open_groups = wb.create_sheet("Открытые Группы")
                                    ws_closed_groups = wb.create_sheet("Закрытые Группы")
                                    write_data(ws_open_channels, openchannels)
                                    write_data(ws_closed_channels, closechannels)
                                    write_data(ws_open_groups, openchats)
                                    write_data(ws_closed_groups, closechats)
                                    wb.save(f"{sessions[i].replace('.session', '')}_about.xlsx")
                                    os.system('cls||clear')
                                    print('Ссылки на чаты добавлены в файл, мой командир')
                                    time.sleep(3)
                                    exit_flag = True
                                    client.disconnect()
                                    break
                                else:
                                    print("Пожалуйста, сделайте свой выбор")
                                    time.sleep(2)
                            except ValueError:
                                print("Пожалуйста, сделайте свой выбор")
                                time.sleep(2)
                else:
                    print("Пожалуйста, выберите существующий аккаунт в диапазоне от 0 до", len(sessions) - 1)
                    time.sleep(2)
            except ValueError:
                print("Пожалуйста, выберите существующий аккаунт в диапазоне от 0 до", len(sessions) - 1)
                time.sleep(2)

def write_data(sheet, data):
    sheet.append(["Название", "Количество участников", "Владелец", "Администратор", "ID", "Ссылка"])
    for item in data:
        owner = " (Владелец)" if item.creator else ""
        admin = " (Администратор)" if item.admin_rights is not None else ""
        usernameadd = f"@{item.username}" if hasattr(item, 'username') and item.username is not None else ""
        sheet.append([item.title, item.participants_count, owner, admin, item.id, usernameadd])
        #sheet.append([item.title, item.participants_count, owner, admin, item.id, f"@{item.username}" if hasattr(item, 'username') else ""])
        
